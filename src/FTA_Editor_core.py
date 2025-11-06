"""
FTA Editor Core Logic
Copyright (c) makkiblog.com - BSD-2 License

This module contains the core business logic for Fault Tree Analysis:
- Data structure management
- Probability calculations
- File I/O operations (JSON, XML, Excel)
- Tree manipulation functions
"""

import json
import xml.etree.ElementTree as ET
import re
import copy
from pathlib import Path
from datetime import datetime


def sanitize_name(s):
    """Sanitize node names by removing extra whitespace"""
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip()


class FTACore:
    """Core logic for Fault Tree Analysis operations"""
    
    def __init__(self):
        """Initialize FTA core with default root node"""
        self.fta_data = {
            "id": "root",
            "name": "RootEvent",
            "type": "Root",
            "logicGate": "",
            "probability": 1.0,  # Changed from 0.0 to 1.0 for better initial display
            "children": [],
            "links": [],  # Added missing links field
            "notes": ""   # Added missing notes field
        }
        self.last_saved_file = None
        
        # Metadata fields with default date
        self.title = "Untitled Analysis"
        self.date = datetime.now().strftime("%Y-%m-%d")
        self.mode = "FTA"  # "FTA" or "ETA"
    
    def get_data(self):
        """Get the current FTA data structure"""
        return self.fta_data
    
    def set_data(self, data):
        """Set the FTA data structure"""
        self.fta_data = data
    
    def get_metadata(self):
        """Get metadata (title, date, mode)"""
        return {
            "title": self.title,
            "date": self.date,
            "mode": self.mode
        }
    
    def set_metadata(self, title=None, date=None, mode=None):
        """Set metadata fields"""
        if title is not None:
            self.title = title
        if date is not None:
            self.date = date
        else:
            # If date is not provided but other metadata is being set, update date to current
            if title is not None or mode is not None:
                self.date = datetime.now().strftime("%Y-%m-%d")
        if mode is not None:
            self.mode = mode
    
    # ========== Node Manipulation Methods ==========
    
    def add_node_to_data(self, parent_id, new_node):
        """Add a new node to the data structure under the specified parent"""
        self._add_node_to_data_recursive(self.fta_data, parent_id, new_node)
    
    def _add_node_to_data_recursive(self, current_node, parent_id, new_node):
        """Recursively add node to data structure"""
        if current_node["id"] == parent_id:
            current_node.setdefault("children", []).append(new_node)
        else:
            for child in current_node.get("children", []):
                self._add_node_to_data_recursive(child, parent_id, new_node)
    
    def update_node(self, node_id, updates):
        """Update a node with new data"""
        node = self.find_node_by_id(node_id)
        if node:
            node.update(updates)
            return True
        return False
    
    def delete_node_from_data(self, node_id):
        """Delete a node from the data structure"""
        self._delete_node_from_data_recursive(self.fta_data, node_id)
    
    def _delete_node_from_data_recursive(self, current_node, node_id):
        """Recursively delete node from data structure"""
        current_node["children"] = [
            child for child in current_node.get("children", [])
            if child.get("id") != node_id
        ]
        for child in current_node.get("children", []):
            self._delete_node_from_data_recursive(child, node_id)
    
    def find_node_by_id(self, node_id):
        """Find and return a node by its ID"""
        return self._find_node_by_id_recursive(self.fta_data, node_id)
    
    def _find_node_by_id_recursive(self, current_node, node_id):
        """Recursively find node by ID"""
        if str(current_node.get("id")) == str(node_id):
            return current_node
        for child in current_node.get("children", []):
            result = self._find_node_by_id_recursive(child, node_id)
            if result:
                return result
        return None
    
    # ========== Probability Calculation Methods ==========
    
    def recalculate_probabilities(self):
        """
        Recalculate probabilities for all nodes based on current mode.
        Delegates to FTA or ETA calculation based on self.mode.
        """
        if self.mode == "ETA":
            self._recalculate_eta_probabilities()
        else:
            self._recalculate_fta_probabilities()
    
    def _recalculate_fta_probabilities(self):
        """
        Recalculate probabilities for FTA (Fault Tree Analysis).
        Works bottom-up: children influence parent probability.
        
        Algorithm:
        1. For leaf nodes: use base probability
        2. For nodes with children:
           - AND gate: product(child_probs)
           - OR gate: 1 - product(1 - child_prob)
        3. Apply AND links: multiply with linked probabilities
        4. Apply OR links: union formula with linked probabilities
        
        Note: When a node has children and a logic gate, the parent's base 
        probability is NOT used. The calculated probability is determined 
        solely by the gate type and children probabilities.
        """
        memo = {}
        visiting = set()

        def get_prob(node):
            nid = node.get("id")
            if nid in memo:
                return memo[nid]
            if nid in visiting:
                # Circular reference detected, use base probability
                val = float(node.get("probability", 1.0))
                memo[nid] = val
                return val

            visiting.add(nid)
            children = node.get("children", []) or []
            
            # Calculate base probability from children
            if not children:
                base = float(node.get("probability", 0.0))
            else:
                child_probs = [get_prob(c) for c in children]
                gate_value = node.get("logicGate", "OR")
                # Normalize: treat None, empty string, or missing as OR
                gate = str(gate_value).strip().upper() if gate_value else "OR"
                
                if gate == "AND":
                    # AND gate: product of children probabilities
                    base = round(self._product(child_probs), 6)
                else:
                    # OR gate: union formula (default)
                    base = round(1 - self._product([1 - p for p in child_probs]), 6)

            # Process links
            links = node.get("links", []) or []
            and_probs = []
            or_probs = []
            
            for l in links:
                tid = l.get("target_id")
                rel = (l.get("relation") or "OR").upper()
                if not tid:
                    continue
                target = self.find_node_by_id(tid)
                if not target:
                    continue
                tp = get_prob(target)
                (and_probs if rel == "AND" else or_probs).append(tp)

            # Apply AND links first
            if and_probs:
                base = round(base * self._product(and_probs), 6)
            
            # Apply OR links second
            if or_probs:
                vals = [base] + or_probs
                base = round(1 - self._product([1 - p for p in vals]), 6)

            memo[nid] = base
            visiting.remove(nid)
            node["calculatedProbability"] = base
            return base

        if isinstance(self.fta_data, dict):
            get_prob(self.fta_data)
    
    def _recalculate_eta_probabilities(self):
        """
        Recalculate probabilities for ETA (Event Tree Analysis).
        Works top-down: parent probability flows down to children.
        
        In ETA mode, the calculated probability of a child is the 
        cumulative product of all parent probabilities down to that child.
        """
        def calc_eta(node, parent_prob=1.0):
            # For ETA, calculated probability is cumulative from parent
            base_prob = float(node.get("probability", 1.0))
            
            # Child's calculated probability is parent's calc prob * child's base prob
            calc_prob = round(parent_prob * base_prob, 6)
            node["calculatedProbability"] = calc_prob
            
            # Recursively calculate for children
            for child in node.get("children", []) or []:
                calc_eta(child, calc_prob)
        
        if isinstance(self.fta_data, dict):
            # Start from root with its base probability
            calc_eta(self.fta_data)
    
    def _product(self, nums):
        """Calculate the product of a list of numbers"""
        result = 1
        for n in nums:
            result *= n
        return result
    
    def get_zero_probability_nodes(self):
        """Return list of node IDs with zero probability"""
        zero_nodes = []
        
        def walk(node):
            nid = str(node.get("id"))
            base_prob = float(node.get("probability", 0.0))
            calc_prob = float(node.get("calculatedProbability", base_prob))
            is_zero = (base_prob == 0.0 or calc_prob == 0.0)
            
            if is_zero:
                zero_nodes.append(nid)
            
            for c in node.get("children", []):
                walk(c)
        
        if isinstance(self.fta_data, dict):
            walk(self.fta_data)
        
        return zero_nodes
    
    # ========== File I/O Methods ==========
    
    def load_from_json(self, file_path):
        """
        Load data from JSON file.
        Supports both new format (with metadata) and legacy format.
        
        Args:
            file_path: Path to the JSON file
            
        Returns:
            tuple: (success: bool, error_message: str or None)
        """
        encodings_to_try = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "cp1252"]
        loaded_data = None
        
        for enc in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.read().strip()
                    # Handle double-wrapped JSON
                    if content.startswith("{{"):
                        content = "{" + content[2:]
                    if content.endswith("}}"):
                        content = content[:-1]
                    loaded_data = json.loads(content)
                break
            except (UnicodeDecodeError, json.JSONDecodeError):
                continue
            except Exception as exc:
                return False, f"Failed to read file: {exc}"
        
        if loaded_data is None:
            return False, "Failed to read file with common encodings"
        
        try:
            if not isinstance(loaded_data, dict):
                return False, "JSON root must be an object"
            
            # Check if new format (with metadata)
            if "tree" in loaded_data:
                # New format with metadata
                self.title = loaded_data.get("title", "Untitled Analysis")
                self.date = loaded_data.get("date", datetime.now().strftime("%Y-%m-%d"))
                self.mode = loaded_data.get("mode", "FTA")
                self.fta_data = loaded_data.get("tree", {})
            else:
                # Legacy format - just tree data
                # Check for FTA wrapper
                if "FTA" in loaded_data:
                    self.fta_data = loaded_data.get("FTA")
                else:
                    self.fta_data = loaded_data
                # Reset metadata to defaults with current date
                self.title = "Untitled Analysis"
                self.date = datetime.now().strftime("%Y-%m-%d")
                self.mode = "FTA"
            
            # Normalize the data structure
            self.fta_data = self._normalize_node(self.fta_data)
            self.last_saved_file = file_path
            self.recalculate_probabilities()
            
            return True, None
            
        except Exception as exc:
            return False, f"Failed to process JSON: {exc}"
    
    def _normalize_node(self, node, parent_id="root", idx=0):
        """Normalize a node's structure and data types"""
        raw = node.get("id")
        node["id"] = str(raw) if raw else f"{parent_id}_{idx}"
        node["name"] = sanitize_name(node.get("name", f"Node_{node['id']}"))
        node["type"] = node.get("type", "Event")
        node["probability"] = float(node.get("probability", 1.0))
        node["logicGate"] = (node.get("logicGate") or "OR").upper()
        node["notes"] = node.get("notes", "") or ""
        
        # Normalize links
        links = node.get("links", []) or []
        for l in links:
            if "target_id" in l:
                l["target_id"] = str(l["target_id"])
            l["relation"] = (l.get("relation") or "OR").upper()
        node["links"] = links
        
        # Normalize children recursively
        children = node.get("children", []) or []
        node["children"] = children
        for i, child in enumerate(children):
            self._normalize_node(child, node["id"], i)
        
        return node
    
    def prepare_export_data(self):
        """Prepare data for export (with metadata and tree data)"""
        export_data = {
            "title": self.title,
            "date": self.date,
            "mode": self.mode,
            "tree": copy.deepcopy(self.fta_data)
        }
        return export_data
    
    def save_to_json(self, file_path=None):
        """
        Save data to JSON file with metadata.
        
        Args:
            file_path: Path to save to. If None, uses last_saved_file.
            
        Returns:
            tuple: (success: bool, error_message: str or None)
        """
        if file_path:
            self.last_saved_file = file_path
        elif not self.last_saved_file:
            return False, "No file path specified"
        
        # Update date to current date on save if it's empty
        if not self.date:
            self.date = datetime.now().strftime("%Y-%m-%d")
        
        try:
            with open(self.last_saved_file, 'w', encoding='utf-8') as f:
                json.dump(self.prepare_export_data(), f, indent=2, ensure_ascii=False)
            return True, None
        except Exception as exc:
            return False, f"Failed to save JSON: {exc}"
    
    def export_to_xml(self, file_path):
        """
        Export FTA data to XML format.
        
        Args:
            file_path: Path to save XML file
            
        Returns:
            tuple: (success: bool, error_message: str or None)
        """
        try:
            root = ET.Element("FaultTree")
            self._build_xml_tree(root, self.fta_data)
            tree = ET.ElementTree(root)
            tree.write(file_path, encoding='utf-8', xml_declaration=True)
            return True, None
        except Exception as exc:
            return False, f"Failed to export XML: {exc}"
    
    def _build_xml_tree(self, parent_elem, node):
        """Recursively build XML tree from FTA data"""
        elem = ET.SubElement(
            parent_elem, "Node",
            name=node.get("name", ""),
            type=node.get("type", ""),
            baseProbability=str(node.get("probability", "")),
            calculatedProbability=str(node.get("calculatedProbability", "")),
            logicGate=node.get("logicGate", "")
        )
        
        if node.get("notes", ""):
            notes_elem = ET.SubElement(elem, "Notes")
            notes_elem.text = node.get("notes", "")
        
        links = node.get("links", [])
        if links:
            links_elem = ET.SubElement(elem, "Links")
            for l in links:
                li = ET.SubElement(
                    links_elem, "Link",
                    target=str(l.get("target_id", "")),
                    relation=l.get("relation", "OR")
                )
                target_node = self.find_node_by_id(l.get("target_id"))
                if target_node:
                    li.text = target_node.get("name", "")
        
        for child in node.get("children", []):
            self._build_xml_tree(elem, child)
    
    def export_to_excel(self, file_path):
        """
        Export FTA data to Excel format with hierarchical column structure.
        Root nodes in column A, their children in column B, grandchildren in C, etc.
        Includes notes and linked nodes information.
        
        Args:
            file_path: Path to save Excel file
            
        Returns:
            tuple: (success: bool, error_message: str or None)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "FTA"
            
            # Track the current row for each depth level
            current_row = 1
            
            # Helper function to write node data in hierarchical format
            def write_node(node, depth, row):
                """Write a node and its children to the worksheet"""
                nonlocal current_row
                
                # Column for this depth level (A=1, B=2, C=3, etc.)
                col = depth + 1
                
                # Create node info string with all details
                name = node.get("name", "")
                node_type = node.get("type", "")
                prob = node.get("probability", "")
                calc_prob = node.get("calculatedProbability", "")
                logic_gate = node.get("logicGate", "")
                notes = node.get("notes", "")
                links = node.get("links", []) or []
                
                # Build the cell value with node information
                cell_value = name
                details = []
                if node_type:
                    details.append(f"Type: {node_type}")
                if prob != "":
                    details.append(f"P: {prob}")
                if calc_prob != "":
                    details.append(f"Calc: {calc_prob}")
                if logic_gate:
                    details.append(f"Gate: {logic_gate}")
                
                if details:
                    cell_value += f"\n({', '.join(details)})"
                
                # Add notes if present
                if notes:
                    cell_value += f"\nNotes: {notes}"
                
                # Add linked nodes information if present
                if links:
                    links_info = []
                    for link in links:
                        target_id = link.get("target_id")
                        relation = link.get("relation", "OR")
                        target_node = self.find_node_by_id(target_id)
                        target_name = target_node.get("name") if target_node else target_id
                        links_info.append(f"{relation}â†’{target_name}")
                    
                    if links_info:
                        cell_value += f"\nLinks: {', '.join(links_info)}"
                
                # Write the cell
                cell = ws.cell(row=row, column=col)
                cell.value = cell_value
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Style based on depth
                colors = ['E6F3FF', 'FFF4E6', 'F0E6FF', 'E6FFE6', 'FFE6F0', 'FFFFE6']
                if depth < len(colors):
                    cell.fill = PatternFill(start_color=colors[depth], 
                                           end_color=colors[depth], 
                                           fill_type='solid')
                
                # Bold for root
                if depth == 0:
                    cell.font = Font(bold=True, size=11)
                
                children = node.get("children", []) or []
                
                if children:
                    # Process children
                    child_start_row = row
                    for i, child in enumerate(children):
                        if i > 0:
                            # Move to next row for siblings
                            current_row += 1
                        write_node(child, depth + 1, current_row)
                else:
                    # Leaf node - just use current row
                    pass
            
            # Start writing from root
            write_node(self.fta_data, 0, current_row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            # Split by newlines and get max line length
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines)
                            max_length = max(max_length, max_line_length)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row heights for better visibility
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 45
            
            wb.save(file_path)
            return True, None
        except Exception as exc:
            return False, f"Failed to export Excel: {exc}"
    
    def get_all_nodes_flat(self):
        """Get a flat list of all nodes with their IDs and names"""
        nodes = []
        
        def collect_nodes(n):
            nodes.append((n.get("id"), n.get("name", "")))
            for c in n.get("children", []):
                collect_nodes(c)
        
        collect_nodes(self.fta_data)
        return nodes
