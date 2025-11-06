#!/usr/bin/env python3
"""
Test script for hierarchical Excel export
Creates a test Excel file to verify the new nested column structure
"""
import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.FTA_Editor_core import FTACore

def test_hierarchical_export():
    """Test the hierarchical Excel export"""
    import tempfile
    import os
    
    print("Testing Hierarchical Excel Export")
    print("=" * 70)
    
    # Create core instance and load sample data
    core = FTACore()
    
    # Find sampleFTA.json in data/examples directory
    sample_data_path = Path(__file__).parent.parent / "data" / "examples" / "sampleFTA.json"
    success, error = core.load_from_json(str(sample_data_path))
    
    if not success:
        print(f"❌ Failed to load sample data: {error}")
        print(f"   Looked for: {sample_data_path}")
        return False
    
    print(f"✓ Loaded sampleFTA.json from {sample_data_path}")
    
    # Recalculate probabilities
    core.recalculate_probabilities()
    print("✓ Calculated probabilities")
    
    # Export to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
        output_file = f.name
    
    try:
        success, error = core.export_to_excel(output_file)
        
        if not success:
            print(f"❌ Failed to export: {error}")
            return False
        
        print(f"✓ Exported to {output_file}")
        
        # Verify the file was created
        if not Path(output_file).exists():
            print("❌ Output file not found")
            return False
        
        # Try to read and verify structure
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            print("\n" + "=" * 70)
            print("Excel Structure Preview:")
            print("=" * 70)
            
            # Show first 15 rows with column positions
            for row_idx in range(1, min(16, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(6, ws.max_column + 1)):  # Show first 5 columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        # Truncate long values
                        val = str(cell.value).split('\n')[0]  # Just show first line
                        if len(val) > 25:
                            val = val[:22] + "..."
                        col_letter = chr(64 + col_idx)  # A, B, C, etc.
                        row_data.append(f"{col_letter}{row_idx}: {val}")
                
                if row_data:
                    print(f"Row {row_idx:2d}: " + " | ".join(row_data))
            
            print("\n" + "=" * 70)
            print(f"Total rows: {ws.max_row}")
            print(f"Total columns: {ws.max_column}")
            print("=" * 70)
            
            # Verify hierarchical structure
            print("\nVerifying hierarchical structure:")
            
            # Root should be in column A (column 1)
            root_cell = ws.cell(row=1, column=1)
            if root_cell.value and "Root Event" in str(root_cell.value):
                print("✓ Root event found in column A, row 1")
            else:
                print(f"❌ Root event not in expected position. Found: {root_cell.value}")
                return False
            
            # Check for children in column B
            found_children_in_b = False
            for row_idx in range(1, ws.max_row + 1):
                cell_b = ws.cell(row=row_idx, column=2)
                if cell_b.value:
                    found_children_in_b = True
                    print(f"✓ Found child in column B, row {row_idx}: {str(cell_b.value).split(chr(10))[0][:30]}")
                    break
            
            if not found_children_in_b:
                print("❌ No children found in column B")
                return False
            
            # Check for grandchildren in column C
            found_grandchildren_in_c = False
            for row_idx in range(1, ws.max_row + 1):
                cell_c = ws.cell(row=row_idx, column=3)
                if cell_c.value:
                    found_grandchildren_in_c = True
                    print(f"✓ Found grandchild in column C, row {row_idx}: {str(cell_c.value).split(chr(10))[0][:30]}")
                    break
            
            if not found_grandchildren_in_c:
                print("❌ No grandchildren found in column C")
                return False
            
            print("\n" + "=" * 70)
            print("✅ ALL TESTS PASSED")
            print("=" * 70)
            print(f"\nGenerated Excel file: {output_file}")
            print("You can open it to verify the hierarchical structure.")
            
            return True
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    finally:
        if os.path.exists(output_file):
            os.unlink(output_file)


if __name__ == "__main__":
    success = test_hierarchical_export()
    sys.exit(0 if success else 1)
